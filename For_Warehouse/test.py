import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import shutil
from datetime import datetime, timedelta

def cash_rest():
    try:
        # Загружаем данные из файлов Excel
        excel_file = r'For_Warehouse\ESP_sklad.xlsx'
        
        # Чтение листа "Склад_остатки"
        rest = pd.read_excel(excel_file, sheet_name='Склад_остатки', skiprows=3, index_col=1)
        
        # Чтение листа "справочник_товаров"
        price = pd.read_excel(excel_file, sheet_name='справочник_товаров', index_col=0)
        
        # Проверяем наличие столбца "Наименование товара" в обеих таблицах
        if 'Наименование товара' not in rest.columns or 'Наименование товара' not in price.columns:
            print("Ошибка: Столбец 'Наименование товара' отсутствует в одном из файлов.")
            return None
        
        # Объединяем таблицы по столбцу "Наименование товара"
        rest = pd.merge(rest, price, how='left', on='Наименование товара')
        
        # Рассчитываем общую стоимость товаров на складе (стоимость * остаток)
        rest['total'] = rest['Стоимость'] * rest['Остаток']
        
        # Рассчитываем сумму всех остатков (общая стоимость)
        total_rest = round(rest['total'].sum(), 2)
        
        # Выводим данные и результат в консоль
        print(rest[['Наименование товара', 'Остаток', 'Стоимость', 'total']])  # Отображаем только важные столбцы
        print(f"Общая стоимость товаров на складе: {total_rest} рублей")
        
        # Возвращаем общую сумму
        return total_rest
    except Exception as e:
        print(f"Ошибка: {e}")
        return None
def rest_of_good(article):
    rest = pd.read_excel(r'For_Warehouse\ESP_sklad.xlsx', skiprows=3, sheet_name='Склад_остатки')
    price = pd.read_excel(r'For_Warehouse\ESP_sklad.xlsx', sheet_name='справочник_товаров')
    rest = pd.merge(rest, price, how='left', on='Наименование товара')
    rest['total'] = rest['Стоимость'] * rest['Остаток']
    article = int(article)
    res = rest.loc[rest['артикул_y'] == article]
    if rest.empty == True:
        s = "Нет товара с таким артикулом"
    elif res['Остаток'].empty:
        s = "Остатоков этого товара нет на складе"
    else:
        s1 = f"{res['Наименование товара']}"
        n = s1[5:(s1.find("\n"))]
        s2 = f"{res['Остаток']}"
        count = s2[5:(s2.find("."))]
        s3 = f"{res['total']}"
        cost = s3[5:(s3.find("\n"))]
        s = f'{n} \t{count} штук\t{cost} рублей'
    return s
def article():
    name = pd.read_excel(r'For_Warehouse\ESP_sklad.xlsx', index_col=1, sheet_name='справочник_товаров')
    s = name['артикул'].to_string()
    return s

def rest_of_controllers():
    # Загружаем данные
    rest = pd.read_excel(r'For_Warehouse\ESP_sklad.xlsx', skiprows=3, sheet_name='Склад_остатки')

    # Фильтруем контроллеры (учитываем регистр)
    res = rest[rest['Наименование товара'].str.contains("Контроллер|контроллер", na=False, regex=True)]

    # Оставляем только нужные столбцы
    res = res[['Наименование товара', 'Остаток']]

    # Группируем и суммируем остатки
    res = res.groupby('Наименование товара', as_index=False).sum()

    # Формируем строку с результатами
    s = "\n".join([
        f"{row['Наименование товара']} {int(row['Остаток'])} штук"
        for _, row in res.iterrows() if not pd.isna(row['Остаток']) and row['Остаток'] > 0
    ])

    return s if s else "Нет остатков контроллеров на складе"

def post_new(name, type, amount):
    date = datetime.today().strftime('%d.%m.%Y')

    # Определение типа операции (приход/отгрузка)
    if type == "shipment":
        new_data = {
            'Дата': [date],
            'Наименование товара': [name],
            'Поступление': [0],
            'Отгрузка': [int(amount)],
            'Комментарий': ['Отгрузка tg-bot']
        }
    else:
        new_data = {
            'Дата': [date],
            'Наименование товара': [name],
            'Поступление': [int(amount)],
            'Отгрузка': [0],
            'Комментарий': ['Поступление tg-bot']
        }

    # Создание DataFrame
    new_data = pd.DataFrame(new_data)

    try:
        # Открытие файла Excel
        wb = load_workbook(r'For_Warehouse\ESP_sklad.xlsx')

        # Проверка наличия листа "Движение_склад"
        if "Движение_склад" not in wb.sheetnames:
            ws = wb.create_sheet("Движение_склад")
        else:
            ws = wb["Движение_склад"]

        # Добавление данных в Excel
        for r in dataframe_to_rows(new_data, header=False, index=False):
            ws.append(r)

        # Сохранение с обработкой ошибки занятости файла
        temp_filename = "temp_table.xlsx"
        wb.save(temp_filename)
        shutil.move(temp_filename, r'For_Warehouse\ESP_sklad.xlsx')

        print("✅ Данные успешно добавлены!")

    except PermissionError:
        print("❌ Ошибка: Закройте 'ESP_sklad.xlsx' и попробуйте снова.")

    except Exception as e:
        print(f"❌ Неизвестная ошибка: {e}")

def shipments(n):
    # Загружаем данные из Excel
    warehouse_move = pd.read_excel(r'For_Warehouse\ESP_sklad.xlsx', sheet_name='Движение_склад')
    price = pd.read_excel(r'For_Warehouse\ESP_sklad.xlsx', sheet_name='справочник_товаров')

    # Преобразуем даты, обрабатывая ошибки
    warehouse_move['Дата'] = pd.to_datetime(warehouse_move['Дата'], dayfirst=True, errors='coerce')

    # Фильтрация по дате
    today = datetime.today()
    if n == 0:
        warehouse_move = warehouse_move[today - warehouse_move['Дата'] <= timedelta(days=7)]
    elif n == 1:
        warehouse_move = warehouse_move[today - warehouse_move['Дата'] <= timedelta(days=30)]

    # Объединение с ценами товаров
    warehouse_move = pd.merge(warehouse_move, price, how='left', on='Наименование товара')
    warehouse_move['Стоимость'] = warehouse_move['Стоимость'].fillna(0)  # Заполняем NaN значением 0
    warehouse_move['sum'] = warehouse_move['Стоимость'] * warehouse_move['Отгрузка']

    # Группировка данных
    res = warehouse_move[['Наименование товара', 'Отгрузка', 'sum']].groupby('Наименование товара', as_index=False).sum()

    # Заполняем NaN нулями
    res['Отгрузка'] = res['Отгрузка'].fillna(0)
    res['sum'] = res['sum'].fillna(0)

    # Формирование итоговой строки
    s = ""
    for _, row in res.iterrows():
        if row['Отгрузка'] > 0:
            s += f"{row['Наименование товара']} {int(row['Отгрузка'])} штук {int(row['sum'])} рублей\n"

    return s if s else "Нет отгрузок за данный период времени"

def delivery(n):
    # Загружаем данные
    warehouse_move = pd.read_excel(r'For_Warehouse\ESP_sklad.xlsx', sheet_name='Движение_склад')
    price = pd.read_excel(r'For_Warehouse\ESP_sklad.xlsx', sheet_name='справочник_товаров')

    # Преобразуем даты (если в Excel был текст, `errors='coerce'` заменит ошибки на NaT)
    warehouse_move['Дата'] = pd.to_datetime(warehouse_move['Дата'], dayfirst=True, errors='coerce')

    # Фильтрация по дате
    today = datetime.today()
    if n == 0:
        warehouse_move = warehouse_move[today - warehouse_move['Дата'] <= timedelta(days=7)]
    elif n == 1:
        warehouse_move = warehouse_move[today - warehouse_move['Дата'] <= timedelta(days=30)]

    # Объединяем с таблицей цен
    warehouse_move = pd.merge(warehouse_move, price, how='left', on='Наименование товара')
    warehouse_move['Стоимость'] = warehouse_move['Стоимость'].fillna(0)  # Заполняем NaN нулями

    # Вычисляем стоимость поступлений
    warehouse_move['sum'] = warehouse_move['Стоимость'] * warehouse_move['Поступление']

    # Группировка данных
    res = warehouse_move[['Наименование товара', 'Поступление', 'sum']].groupby('Наименование товара', as_index=False).sum()

    # Заполняем NaN нулями
    res['Поступление'] = res['Поступление'].fillna(0)
    res['sum'] = res['sum'].fillna(0)

    # Формируем строку для вывода
    s = ""
    for _, row in res.iterrows():
        if row['Поступление'] > 0:
            s += f"{row['Наименование товара']} {int(row['Поступление'])} штук {int(row['sum'])} рублей\n"

    return s if s else "Нет поступлений за данный период времени"

def post_new(name, operation_type, amount):
    date = datetime.today().strftime('%d.%m.%Y')

    # Создаем новую строку с данными
    new_data = {
        'Дата': [date],
        'Наименование товара': [name],
        'Поступление': [int(amount) if operation_type != "shipment" else 0],
        'Отгрузка': [int(amount) if operation_type == "shipment" else 0],
        'Комментарий': [f'{"Отгрузка" if operation_type == "shipment" else "Поступление"} tg-bot']
    }
    
    new_data = pd.DataFrame(new_data)

    # Загружаем существующий файл Excel
    try:
        wb = load_workbook(filename="table.xlsx", data_only=True)
    except FileNotFoundError:
        return "Ошибка: файл table.xlsx не найден."

    # Проверяем, есть ли лист "Движение_склад"
    if "Движение_склад" not in wb.sheetnames:
        return "Ошибка: лист 'Движение_склад' не найден в файле."

    ws = wb["Движение_склад"]

    # Добавляем новую строку
    for row in new_data.itertuples(index=False, name=None):
        ws.append(row)

    # Сохраняем файл
    wb.save("table.xlsx")
    
    return "✅ Данные успешно добавлены!"

def is_correct_article(article):
    try:
        # Загружаем Excel-файл
        goods = pd.read_excel(r'For_Warehouse\ESP_sklad.xlsx', sheet_name='справочник_товаров')

        # Проверяем, есть ли нужные столбцы
        if 'артикул' not in goods.columns or 'Наименование товара' not in goods.columns:
            return "Ошибка: Нет нужных столбцов в файле."

        # Ищем товар по артикулу
        goods = goods.loc[goods['артикул'] == int(article)]

        # Если товара нет, возвращаем 0
        if goods.empty:
            return 0

        # Возвращаем название товара
        return goods.iloc[0]['Наименование товара']

    except (ValueError, FileNotFoundError, KeyError):
        return 0

