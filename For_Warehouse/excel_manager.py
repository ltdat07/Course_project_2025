import os
import logging
from datetime import datetime
from openpyxl import load_workbook
from filelock import FileLock
import pandas as pd
import asyncio

class ExcelManager:
    def __init__(self, file_path: str, cache_duration: int = 60):
        """
        :param file_path: Путь к Excel файлу.
        :param cache_duration: Время кэширования в секундах (по умолчанию 60 секунд).
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel файл не найден: {file_path}")
        self.file_path = file_path
        self.cache_duration = cache_duration
        self._workbook = None
        self.last_loaded = None

    async def load_workbook_async(self):
        """
        Асинхронно загружает рабочую книгу с учетом кэширования.
        """
        now = datetime.now()
        if self._workbook and self.last_loaded and (now - self.last_loaded).seconds < self.cache_duration:
            return self._workbook
        lock = FileLock(self.file_path + ".lock")
        with lock:
            # Выполняем загрузку в отдельном потоке
            self._workbook = await asyncio.to_thread(load_workbook, self.file_path)
            self.last_loaded = now
        logging.info("Excel файл загружен заново.")
        return self._workbook

    async def save_workbook_async(self):
        """
        Асинхронно сохраняет текущую рабочую книгу с блокировкой.
        """
        lock = FileLock(self.file_path + ".lock")
        with lock:
            await asyncio.to_thread(self._workbook.save, self.file_path)
            self.last_loaded = datetime.now()
        logging.info("Excel файл сохранен.")

    def get_sheet(self, sheet_name: str):
        """
        Возвращает рабочий лист по его имени.
        """
        wb = self._workbook
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Лист '{sheet_name}' не найден в книге")
        return wb[sheet_name]

    def read_sheet_as_df(self, sheet_name: str, **kwargs) -> pd.DataFrame:
        """
        Считывает лист Excel в DataFrame.
        """
        return pd.read_excel(self.file_path, sheet_name=sheet_name, **kwargs)
