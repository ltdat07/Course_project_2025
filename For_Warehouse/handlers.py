import pandas as pd
import os
import logging
from datetime import datetime, timedelta
from openpyxl.styles import Alignment
from filelock import FileLock
import config
from excel_manager import ExcelManager
from reports import generate_stock_report  # Импортируем новую универсальную функцию
from utils import format_currency
from decorators import log_exceptions
import asyncio

# Глобальное состояние для каждого пользователя
user_state = {}

def resetUserState(chat_id):
    user_state[chat_id] = {
        "operation": None,
        "step": None,
        "product_name": None,
        "quantity": None,
        "comment": None,
        "employee": None
    }
    logging.info(f"Состояние для пользователя {chat_id} сброшено.")

def logUserStateChange(chat_id, new_state):
    logging.info(f"Пользователь {chat_id} перешёл к состоянию: {new_state}")

# Глобальный экземпляр ExcelManager
excel_manager = ExcelManager(config.settings.EXCEL_FILE)

@log_exceptions
def getTotalStockValue():
    inventory_df = excel_manager.read_sheet_as_df(config.settings.SHEET_REST, skiprows=3, index_col=1)
    price_df = excel_manager.read_sheet_as_df(config.settings.SHEET_PRICE, index_col=0)
    if 'Наименование товара' not in inventory_df.columns or 'Наименование товара' not in price_df.columns:
        logging.error("Столбец 'Наименование товара' отсутствует.")
        return None
    inventory_df = pd.merge(inventory_df, price_df, how='left', on='Наименование товара')
    inventory_df['total'] = inventory_df['Стоимость'] * inventory_df['Остаток']
    total_inventory = round(inventory_df['total'].sum(), 2)
    formatted_total = f"{total_inventory:,.2f}".replace(",", " ")
    result = (
        "💰 <b>Общий остаток на складе</b>\n"
        "────────────────────────────\n"
        f"🔢 <b>Сумма:</b> <code>{formatted_total} руб.</code>"
    )
    return result

@log_exceptions
def getProductStock(article):
    # Читаем данные из листа остатков и прайс-листа
    inventory_df = excel_manager.read_sheet_as_df(config.settings.SHEET_REST, skiprows=3)
    price_df = excel_manager.read_sheet_as_df(config.settings.SHEET_PRICE)
    
    # Задаём явные суффиксы, чтобы получить колонку 'Артикул' из прайс-листа под именем "Артикул_price"
    inventory_df = pd.merge(
        inventory_df, 
        price_df, 
        how='left', 
        on='Наименование товара', 
        suffixes=('_rest', '_price')
    )
    
    inventory_df['total'] = inventory_df['Стоимость'] * inventory_df['Остаток']
    article = str(article).strip().lstrip("0")
    
    # Приводим значения столбца из прайс-листа к корректному виду
    inventory_df['Артикул_price'] = inventory_df['Артикул_price'].astype(str).str.strip().str.lstrip("0")
    
    product_df = inventory_df.loc[inventory_df['Артикул_price'] == article]
    if product_df.empty:
        return "Нет товара с таким артикулом"
    if product_df['Остаток'].empty:
        return "Остатков этого товара нет на складе"
    prod_name = product_df.iloc[0]['Наименование товара']
    prod_quantity = int(product_df.iloc[0]['Остаток'])
    prod_total = round(product_df.iloc[0]['total'], 2)
    return (prod_name, prod_quantity, prod_total)

@log_exceptions
def validateProductArticle(article):
    goods_df = excel_manager.read_sheet_as_df(config.settings.SHEET_PRICE)
    # Убираем лишние пробелы в названиях столбцов, не приводя их к нижнему регистру 
    goods_df.columns = goods_df.columns.str.strip()
    if 'Артикул' not in goods_df.columns or 'Наименование товара' not in goods_df.columns:
        logging.error("Нет нужных столбцов в файле.")
        return "Ошибка: Нет нужных столбцов в файле."
    goods_df['Артикул'] = goods_df['Артикул'].astype(str).str.strip()
    article_str = str(article).strip().lstrip("0")
    logging.info(f"Ищем артикул: '{article_str}'")
    matching_goods = goods_df.loc[goods_df['Артикул'] == article_str]
    if matching_goods.empty:
        logging.info(f"Артикул '{article_str}' не найден.")
        return 0
    product_row = matching_goods.iloc[0]
    logging.info(f"Найден артикул: {product_row['Артикул']}, товар: {product_row['Наименование товара']}")
    return (product_row['Артикул'], product_row['Наименование товара'])

def listProductArticles():
    try:
        price_data = excel_manager.read_sheet_as_df(config.settings.SHEET_PRICE)
        price_data.columns = price_data.columns.str.strip()
        price_data = price_data[['Артикул', 'Наименование товара']].dropna()
        articles = []
        for _, row in price_data.iterrows():
            art_code = str(row['Артикул']).strip()
            articles.append(f"{art_code}: {row['Наименование товара']}")
        return "\n".join(articles)
    except Exception as err:
        logging.error(f"Ошибка в listProductArticles: {err}")
        return f"Ошибка: {err}"

def listControllerStocks():
    try:
        inventory_df = excel_manager.read_sheet_as_df(config.settings.SHEET_REST, skiprows=3)
        controllers_df = inventory_df[inventory_df['Наименование товара'].str.contains("Контроллер|контроллер", na=False, regex=True)]
        controllers_df = controllers_df[['Артикул', 'Наименование товара', 'Остаток']]
        grouped_controllers = controllers_df.groupby(['Артикул', 'Наименование товара'], as_index=False).sum()
        if grouped_controllers.empty or grouped_controllers['Остаток'].sum() == 0:
            return "❌ <i>Нет остатков контроллеров на складе</i>"
        formatted_list = "\n".join([
            f"🔹 <b>{row['Наименование товара']}</b> (Артикул: {row['Артикул']}) — <code>{int(row['Остаток'])} шт.</code>"
            for _, row in grouped_controllers.iterrows()
        ])
        return f"📦 <b>Остаток контроллеров на складе:</b>\n\n{formatted_list}"
    except Exception as err:
        logging.error(f"Ошибка в listControllerStocks: {err}")
        return f"🚨 <i>Ошибка:</i> {err}"

def findFirstEmptyRow(worksheet, start_row=2):
    for row_idx in range(start_row, worksheet.max_row + 1):
        row_cells = worksheet[row_idx]
        if all(cell.value is None or str(cell.value).strip() == "" for cell in row_cells):
            return row_idx
    return worksheet.max_row + 1

@log_exceptions
async def recordWarehouseOperation(name, operation_type, amount, user_comment="", employee=""):
    """
    Асинхронно записывает операцию (поступление или отгрузку) в лист "Движение_склад" 
    и обновляет данные в листе "Склад_остатки".
    """
    current_date = datetime.today()
    operation_data = {
        'Дата': [current_date],
        'Наименование товара': [name],
        'Поступление': [int(amount) if operation_type == "delivery" else 0],
        'Отгрузка': [int(amount) if operation_type == "shipment" else 0],
        'Комментарий': [user_comment],
        'Сотрудник': [employee]
    }
    operation_df = pd.DataFrame(operation_data)
    try:
        wb = await excel_manager.load_workbook_async()
    except Exception as err:
        logging.error(f"Ошибка при открытии файла: {err}")
        return f"❌ Ошибка при открытии файла: {err}"
    
    move_sheet = config.settings.SHEET_MOVE
    if move_sheet not in wb.sheetnames:
        logging.error(f"Лист '{move_sheet}' не найден.")
        return f"❌ Ошибка: Лист '{move_sheet}' не найден в файле."
    ws_move = wb[move_sheet]
    empty_row = findFirstEmptyRow(ws_move, start_row=2)
    for col_idx, col_name in enumerate(operation_df.columns, start=1):
        if col_name == "Дата":
            cell = ws_move.cell(row=empty_row, column=col_idx, value=operation_df.iloc[0][col_name])
            cell.number_format = "DD.MM.YY"
            cell.alignment = Alignment(horizontal='right')
        else:
            ws_move.cell(row=empty_row, column=col_idx, value=operation_df.iloc[0][col_name])
    
    rest_sheet = config.settings.SHEET_REST
    if rest_sheet not in wb.sheetnames:
        logging.error(f"Лист '{rest_sheet}' не найден.")
        return f"❌ Ошибка: Лист '{rest_sheet}' не найден в файле."
    ws_rest = wb[rest_sheet]
    header_row = 4
    header_map = {}
    for cell in ws_rest[header_row]:
        if cell.value and isinstance(cell.value, str):
            header_map[cell.value.strip()] = cell.column
    for req_col in ["Наименование товара", "Поступление", "Отгрузка", "Остаток"]:
        if req_col not in header_map:
            logging.error(f"Столбец '{req_col}' не найден.")
            return f"❌ Ошибка: Не найден столбец '{req_col}' в листе 'Склад_остатки'."
    col_product = header_map["Наименование товара"]
    col_delivery = header_map["Поступление"]
    col_shipment = header_map["Отгрузка"]
    col_balance = header_map["Остаток"]
    product_exists = False
    for row_idx in range(header_row + 1, ws_rest.max_row + 1):
        cell_val = ws_rest.cell(row=row_idx, column=col_product).value
        if cell_val and str(cell_val).strip().lower() == str(name).strip().lower():
            product_exists = True
            curr_delivery = ws_rest.cell(row=row_idx, column=col_delivery).value or 0
            curr_shipment = ws_rest.cell(row=row_idx, column=col_shipment).value or 0
            curr_balance = ws_rest.cell(row=row_idx, column=col_balance).value or 0
            try:
                curr_delivery = int(curr_delivery)
            except:
                curr_delivery = 0
            try:
                curr_shipment = int(curr_shipment)
            except:
                curr_shipment = 0
            try:
                curr_balance = int(curr_balance)
            except:
                curr_balance = 0
            if operation_type == "delivery":
                new_delivery = curr_delivery + int(amount)
                new_balance = curr_balance + int(amount)
                ws_rest.cell(row=row_idx, column=col_delivery, value=new_delivery)
            elif operation_type == "shipment":
                new_shipment = curr_shipment + int(amount)
                new_balance = curr_balance - int(amount)
                ws_rest.cell(row=row_idx, column=col_shipment, value=new_shipment)
            else:
                return "❌ Ошибка: Неверный тип операции."
            ws_rest.cell(row=row_idx, column=col_balance, value=new_balance)
            break
    if not product_exists:
        logging.error(f"Товар '{name}' не найден.")
        return f"❌ Ошибка: Товар '{name}' не найден в листе 'Склад_остатки'."
    try:
        await excel_manager.save_workbook_async()
        logging.info(f"Операция для '{name}' успешно записана.")
        return "✅ Данные успешно добавлены и данные по складу обновлены!"
    except PermissionError:
        return "❌ Ошибка: Файл открыт в Excel. Закройте его и попробуйте снова."
    except Exception as err:
        logging.error(f"Ошибка при записи: {err}")
        return f"❌ Ошибка при записи: {err}"

# Обновлённые функции для вызова отчётов

def listDeliveriesLastNDays(n):
    return generate_stock_report(n, "delivery")

def listShipmentsLastNDays(n):
    return generate_stock_report(n, "shipment")
